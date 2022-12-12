import requests
import os
from zipfile import ZipFile
from openpyxl import Workbook
from PyPDF2 import PdfReader


# Получить родительскую директорию:


# parent = os.pardir
# path = os.getcwd()
# print(path)
# pathpar = os.path.join(path, parent)
# print(pathpar)


# Cкачивание файла pdf


resourses_pdf_dir = (os.path.abspath(os.path.join(os.path.dirname(__file__), '..', r'resourses\pdf_test.pdf')))
pdf_url = 'https://filesamples.com/samples/document/pdf/sample1.pdf'
response = requests.get(pdf_url, stream=True)
with open(resourses_pdf_dir, 'wb') as pdf_file:
    pdf_file.write(response.content)


# Создание файла xlsx


resourses_xlsx_dir = (os.path.abspath(os.path.join(os.path.dirname(__file__), '..', r'resourses\xlsx_test.xlsx')))
book = Workbook()
sheet = book.active
sheet['A1'] = 'Hello'
sheet['A2'] = 'World'
sheet['B1'] = 1
sheet['B2'] = 2
book.save(resourses_xlsx_dir)


# Cкачивание файла xlsx


# xlsx_url = 'https://freetestdata.com/wp-content/uploads/2021/09/Free_Test_Data_100KB_XLSX.xlsx'
# response = requests.get(xlsx_url, stream = True)
#
# with open(resourses_xlsx_dir, 'wb') as xlsx_file:
#     xlsx_file.write(response.content)


# Cкачивание файла csv


resourses_csv_dir = (os.path.abspath(os.path.join(os.path.dirname(__file__), '..', r'resourses\csv_test.csv')))
csv_url = 'https://people.sc.fsu.edu/~jburkardt/data/csv/addresses.csv'
response = requests.get(csv_url, stream=True)
with open(resourses_csv_dir, 'wb') as csv_file:
    csv_file.write(response.content)


# Создание и запись файлов в zip


resourses_zip_dir = (os.path.abspath(os.path.join(os.path.dirname(__file__), '..', r'resourses\zip_test.zip')))


def test_create_zip():
    with ZipFile(resourses_zip_dir, 'w') as zip_create_file:
        zip_create_file.write(resourses_csv_dir, arcname='csv_test.csv')
        print('\nCSV file archive success!')
        zip_create_file.write(resourses_pdf_dir, arcname='pdf_test.pdf')
        print('PDF file archive success!')
        zip_create_file.write(resourses_xlsx_dir, arcname='xlsx_test.xlsx')
        print('XLSX file archive success!')


# Чтение и проверка содержимого заархивированных файлов в zip


def test_read_pdf_file():
    with ZipFile(resourses_zip_dir, "r") as zip_read_pdf_file:
        print('\nReading & Asserting pdf file...')
        zip_read_pdf_file.extract('pdf_test.pdf')
        pdf_content = PdfReader('pdf_test.pdf').pages[0].extract_text()
        os.remove('pdf_test.pdf')
        assert pdf_content.__contains__('Instructions for')
        print('Done!')


def test_read_csv_file():
    with ZipFile(resourses_zip_dir, "r") as zip_read_csv_file:
        print('\nReading & Asserting csv file...')
        csv_content = str(zip_read_csv_file.read('csv_test.csv'))
        assert csv_content.__contains__('Doe')
        print('Done!')


def test_read_xlsx_file():
    with ZipFile(resourses_zip_dir, "r") as zip_read_xlsx_file:
        print('\nReading & Asserting xlsx file...')
        str(zip_read_xlsx_file.read('xlsx_test.xlsx'))
        assert sheet.cell(row=1, column=1).value == 'Hello'
        assert sheet.cell(row=2, column=1).value == 'World'
        assert sheet.cell(row=1, column=2).value == 1
        assert sheet.cell(row=2, column=2).value == 2
        print('Done!')
